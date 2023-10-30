import os
import zipfile

import xlrd
from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from utils.constants import RESOURCES_DIR, ARCHIVE_PATH, PDF_FILE, TXT_FILE, XLS_FILE, XLSX_FILE, LIST_OF_FILES


def test_archive_exists(create_archive):
    assert os.path.exists(ARCHIVE_PATH)


def test_check_files_in_archive(create_archive):
    archive_zip_list = ZipFile(ARCHIVE_PATH).namelist()
    for zipped_file in archive_zip_list:
        assert zipped_file in LIST_OF_FILES


def test_archive_contents(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH, 'r') as archive:
        for filename in archive.namelist():
            with archive.open(filename) as file_in_archive:
                if filename.endswith('.pdf'):
                    archived_pdf = PdfReader(file_in_archive)
                    expected_pdf_text = archived_pdf.pages[0].extract_text()
                    original_pdf = PdfReader(os.path.join(RESOURCES_DIR, PDF_FILE))
                    original_txt_content = original_pdf.pages[0].extract_text()
                    assert expected_pdf_text == original_txt_content
                elif filename.endswith('.xls'):
                    book = xlrd.open_workbook(file_contents=file_in_archive.read())
                    sheet = book.sheet_by_index(0)
                    archived_xls_cell_content = str(sheet.cell_value(2, 2)).strip()
                    original_book = xlrd.open_workbook(os.path.join(RESOURCES_DIR, XLS_FILE))
                    original_sheet = original_book.sheet_by_index(0)
                    original_xls_cell_content = (original_sheet.cell_value(2, 2)).strip()
                    assert archived_xls_cell_content == original_xls_cell_content
                elif filename.endswith('.xlsx'):
                    wb = load_workbook(file_in_archive)
                    sheet = wb.active
                    archived_xlsx_cell_content = sheet['B4'].value
                    wb_orig = load_workbook(os.path.join(RESOURCES_DIR, XLSX_FILE))
                    sheet_orig = wb_orig.active
                    original_xlsx_cell_content = sheet_orig['B4'].value
                    assert archived_xlsx_cell_content == original_xlsx_cell_content
                else:
                    archived_txt_content = file_in_archive.read().decode('utf-8').strip()
                    with open(os.path.join(RESOURCES_DIR, TXT_FILE), 'r') as original_file:
                        original_txt_content = original_file.read().strip()
                    assert archived_txt_content == original_txt_content
